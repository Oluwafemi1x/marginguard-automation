import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
OUT.mkdir(exist_ok=True)


def export_csv(results: list[dict], name: str = "marginguard-report.csv") -> Path:
    path = OUT / name
    if not results:
        path.write_text("", encoding="utf-8")
        return path

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=results[0].keys())
        writer.writeheader()
        writer.writerows(results)
    return path


def export_excel(results: list[dict], name: str = "marginguard-report.xlsx") -> Path:
    path = OUT / name
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Competitive Intelligence"

    headers = [
        "SKU",
        "Product",
        "Competitor",
        "Our Price",
        "Competitor Price",
        "Gap %",
        "Stock",
        "Severity",
        "Opportunity",
        "Recommendation",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="151A2F")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")

    for result in results:
        sheet.append(
            [
                result["sku"],
                result["product_name"],
                result["competitor"],
                result["our_price"],
                result["competitor_price"],
                result["gap_percent"],
                result["stock"],
                result["severity"],
                result["opportunity"],
                result["recommendation"],
            ]
        )

    widths = [14, 28, 18, 14, 18, 12, 12, 14, 16, 58]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)
    return path
